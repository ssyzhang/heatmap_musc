import os
import sys
import numpy as np
from pyparsing import original_text_for
from sympy import im
import torch
import torch.nn.functional as F
sys.path.append('./models/backbone')

import datasets.mvtec as mvtec
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
import datasets.visa as visa
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
import datasets.btad as btad
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad
import datasets.cvlab as cvlab

import models.backbone.open_clip as open_clip
import models.backbone._backbones as _backbones
from models.modules._LNAMD import LNAMD
from models.modules._MSM import MSM
from models.modules._RsCIN import RsCIN
from utils.metrics import compute_metrics
from openpyxl import Workbook
from tqdm import tqdm
import pickle
import time
import cv2

import warnings
warnings.filterwarnings("ignore")


class MuSc():
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        # self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")

        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
            else:
                self.categories = [self.categories]

        self.model_name = cfg['models']['backbone_name']
        self.image_size = cfg['datasets']['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.pretrained = cfg['models']['pretrained']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        self.divide_num = cfg['datasets']['divide_num']
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name, 'imagesize{}'.format(self.image_size))
        os.makedirs(self.output_dir, exist_ok=True)
        self.load_backbone()


    def load_backbone(self):
        if 'dino' in self.model_name:
            # dino or dino_v2
            self.dino_model = _backbones.load(self.model_name)
            self.dino_model.to(self.device)
            self.preprocess = None
        else:
            # clip ,下方函数返回return model, preprocess_train, preprocess_val
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(self.model_name, self.image_size, pretrained=self.pretrained)
            self.clip_model.to(self.device)


    def load_datasets(self, category, divide_num=1, divide_iter=0):
        # dataloader
        if self.dataset == 'visa':
            test_dataset = visa.VisaDataset(source=self.path, split=visa.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=mvtec.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'btad':
            test_dataset = btad.BTADDataset(source=self.path, split=btad.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'cvlab':
            test_dataset = cvlab.CVLabDataset(source=self.path, split=cvlab.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        return test_dataset
    def visualization(self, image_path_list, gt_list, pr_px, category):
        def normalization01(img):
            # 加上防报错机制：防止最大最小值一样时除以0报错
            if img.max() == img.min():
                return img - img.min()
            return (img - img.min()) / (img.max() - img.min())

        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                # anomaly_type = path.split('/')[-2]
                anomaly_type = 'anomaly'
                img_name = path.split('/')[-1].split('\\')[-1]
                
                mask_name = img_name.replace(".png", "_mask.png")
                # 统一的绝对路径
                mask_path = os.path.join("e:/cvlab/data0312/PF/mask", mask_name)
                origin_path = os.path.join("e:/cvlab/data0312/PF/img", img_name)

                if anomaly_type not in['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_dir = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_dir, exist_ok=True)
                    save_file = os.path.join(save_dir, img_name)
                    np_save_path=save_file.replace('.png','.npy')
                    # 1. 生成预测热力图 (3通道 BGR 图像)
                    anomaly_map = pr_px[i].squeeze()
                    np.save(np_save_path, anomaly_map)
                    # anomaly_map = normalization01(anomaly_map) * 255
                    # anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)

                    # # ----------------- 新增：读取原图和Mask并三拼接 -----------------
                    # # 2. 读取原图和掩码图
                    # origin_img = cv2.imread(origin_path)
                    # mask_img = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
                    
                    # # 获取热力图的尺寸作为基准 (高, 宽)
                    # h, w = anomaly_map.shape[:2]
                    
                    # # 准备一个空列表，用来按顺序存放要拼接的图片
                    # vis_list =[]
                    
                    # # 【处理 1：原图】
                    # if origin_img is not None:
                    #     # OpenCV 默认以 BGR 3通道读取，无需转通道，只需检查尺寸是否一致
                    #     if origin_img.shape[:2] != (h, w):
                    #         origin_img = cv2.resize(origin_img, (w, h))
                    #     vis_list.append(origin_img)
                    # else:
                    #     print(f"警告: 找不到原始图片 {origin_path}")

                    # # 【处理 2：真实掩码图】
                    # if mask_img is not None:
                    #     # 0/1二值图转为0-255
                    #     if mask_img.max() <= 1:
                    #         mask_img = mask_img * 255
                    #     # 单通道灰度转为3通道BGR，以便和其他图拼接
                    #     mask_3c = cv2.cvtColor(mask_img.astype(np.uint8), cv2.COLOR_GRAY2BGR)
                    #     if mask_3c.shape[:2] != (h, w):
                    #         mask_3c = cv2.resize(mask_3c, (w, h))
                    #     vis_list.append(mask_3c)
                    # else:
                    #     print(f"警告: 找不到掩码文件 {mask_path}")
                        
                    # # 【处理 3：预测热力图】
                    # vis_list.append(anomaly_map)
                    
                    # # 3. 水平拼接：[ 原图 | 掩码图 | 热力图 ]
                    # # tuple(vis_list) 会把列表里所有存在的图片从左到右拼起来
                    # final_vis = np.hstack(tuple(vis_list))
                    # # -----------------------------------------------------

                    # # 4. 保存最终的三拼接图片
                    # cv2.imwrite(save_file, final_vis)


#以上部分修改
    # def visualization(self, image_path_list, gt_list, pr_px, category):
    #     def normalization01(img):
    #         # 加上防报错机制：防止最大最小值一样时除以0报错
    #         if img.max() == img.min():
    #             return img - img.min()
    #         return (img - img.min()) / (img.max() - img.min())

    #     if self.vis_type == 'single_norm':
    #         # normalized per image
    #         for i, path in enumerate(image_path_list):
    #             # anomaly_type = path.split('/')[-2]
    #             anomaly_type = 'anomaly'
    #             img_name = path.split('/')[-1].split('\\')[-1]
    #             # print(img_name) # PF231.png
                
    #             mask_name = img_name.replace(".png", "_mask.png")
    #             # 【修复Bug】: 强烈建议使用 os.path.join，防止漏掉路径中间的斜杠 '/'
    #             mask_path = os.path.join("e:/cvlab/data0312/PF/mask", mask_name)
    #             origin_path= os.path.join("e:/cvlab/data0312/PF/img", img_name)

    #             if anomaly_type not in['good', 'Normal', 'ok'] and gt_list[i] != 0:
    #                 save_dir = os.path.join(self.output_dir, category, anomaly_type)
    #                 os.makedirs(save_dir, exist_ok=True)
    #                 save_file = os.path.join(save_dir, img_name)
                    
    #                 # 1. 生成热力图 (3通道 BGR 图像)
    #                 anomaly_map = pr_px[i].squeeze()
    #                 anomaly_map = normalization01(anomaly_map) * 255
    #                 anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)

    #                 # ----------------- 新增：读取Mask并拼接 -----------------
    #                 # 2. 读取真实的 mask 掩码图 (以灰度图模式读取)
    #                 mask_img = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
                    
    #                 if mask_img is not None:
    #                     # 掩码如果是0和1的值，为了可视化需要乘255变成黑白图 (如果已经是0-255则不需要)
    #                     if mask_img.max() <= 1:
    #                         mask_img = mask_img * 255
                            
    #                     # 热力图是3通道，掩码是1通道灰度图，必须把掩码转成3通道才能拼接
    #                     mask_3c = cv2.cvtColor(mask_img.astype(np.uint8), cv2.COLOR_GRAY2BGR)
                        
    #                     # 安全检查：确保掩码的尺寸和热力图的尺寸完全一致 (宽, 高)
    #                     h, w = anomaly_map.shape[:2]
    #                     if mask_3c.shape[:2] != (h, w):
    #                         mask_3c = cv2.resize(mask_3c, (w, h))
                            
    #                     # 3. 水平拼接：左边是Mask，右边是热力图
    #                     final_vis = np.hstack((mask_3c, anomaly_map))
    #                 else:
    #                     print(f"警告: 找不到掩码文件 {mask_path}，将只保存热力图。")
    #                     final_vis = anomaly_map
    #                 # -----------------------------------------------------

    #                 # 4. 保存最终的拼接图片
    #                 cv2.imwrite(save_file, final_vis)
    # def visualization(self, image_path_list, gt_list, pr_px, category):
    #     def normalization01(img):
    #         return (img - img.min()) / (img.max() - img.min())
    #     if self.vis_type == 'single_norm':
    #         # normalized per image
    #         for i, path in enumerate(image_path_list):
    #             # anomaly_type = path.split('/')[-2]
    #             anomaly_type='anomaly'
    #             img_name = path.split('/')[-1].split('\\')[-1]
    #             # print(img_name)#PF231.png
    #             mask_name=img_name.replace(".png", "_mask.png")
    #             mask_path = "e:/cvlab/data0312/PF/mask" + mask_name
    #             if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
    #                 save_path = os.path.join(self.output_dir, category, anomaly_type)
    #                 os.makedirs(save_path, exist_ok=True)
    #                 save_path = os.path.join(save_path, img_name)
    #                 anomaly_map = pr_px[i].squeeze()
    #                 anomaly_map = normalization01(anomaly_map)*255
    #                 anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
    #                                # ----------------- 新增：拼接真实掩码 
    #                 # if gt_mask_list is not None:
    #                 #     mask = gt_mask_list[i].squeeze()
    #                 #     if mask.max() <= 1.0 and mask.max() > 0:
    #                 #         mask = mask * 255
                            
    #                 #     mask_3c = cv2.cvtColor(mask.astype(np.uint8), cv2.COLOR_GRAY2BGR)
                        
    #                 #     if mask_3c.shape[:2] != anomaly_map.shape[:2]:
    #                 #         mask_3c = cv2.resize(mask_3c, (anomaly_map.shape[1], anomaly_map.shape[0]))
                            
    #                 #     final_vis = np.hstack((mask_3c, anomaly_map))
    #                 # else:
    #                 #     final_vis = anomaly_map
    #             # -----------------------------------------------------
    #                 cv2.imwrite(save_path, anomaly_map)
    #     else:
    #         # normalized all image
    #         pr_px = normalization01(pr_px)
    #         for i, path in enumerate(image_path_list):
    #             anomaly_type = path.split('/')[-2]
    #             img_name = path.split('/')[-1]
    #             save_path = os.path.join(self.output_dir, category, anomaly_type)
    #             os.makedirs(save_path, exist_ok=True)
    #             save_path = os.path.join(save_path, img_name)
    #             anomaly_map = pr_px[i].squeeze()
    #             anomaly_map *= 255
    #             anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
    #             cv2.imwrite(save_path, anomaly_map)


    def make_category_data(self, category):
        print(category)

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        start_time_all = time.time()
        dataset_num = 0
        for divide_iter in range(divide_num):
            test_dataset = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter)
            test_dataloader = torch.utils.data.DataLoader(
                test_dataset,
                batch_size=self.batch_size,
                shuffle=False,
                num_workers=0,
                pin_memory=True,
            )
            
            # extract features
            patch_tokens_list = []
            subset_num = len(test_dataset)
            dataset_num += subset_num
            start_time = time.time()
            for image_info in tqdm(test_dataloader):
            # for image_info in test_dataloader:
                if isinstance(image_info, dict):
                    image = image_info["image"]
                    image_path_list.extend(image_info["image_path"])
                    img_masks.append(image_info["mask"])
                    gt_list.extend(list(image_info["is_anomaly"].numpy()))
                with torch.no_grad(), torch.cuda.amp.autocast():
                    input_image = image.to(torch.float).to(self.device)
                    if 'dinov2' in self.model_name:
                        patch_tokens = self.dino_model.get_intermediate_layers(x=input_image, n=[l-1 for l in self.features_list], return_class_token=False)
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                        fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                        patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                    elif 'dino' in self.model_name:
                        patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens_all[l-1].cpu() for l in self.features_list]
                    else: # clip
                        image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)
                        image_features /= image_features.norm(dim=-1, keepdim=True)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                image_features = [image_features[bi].squeeze().cpu().numpy() for bi in range(image_features.shape[0])]#这里变为了列表，列表中有b_s个数组
                class_tokens.extend(image_features)
                patch_tokens_list.append(patch_tokens)  # (B, L+1, C)
            end_time = time.time()#cl_tokens83长的768维数组列表,p_t_l21长[bs,1370,1024]
            print('extract time: {}ms per image'.format((end_time-start_time)*1000/subset_num))#处理每张图片用了多少毫秒
            
            # LNAMD
            feature_dim = patch_tokens_list[0][0].shape[-1]
            anomaly_maps_r = torch.tensor([]).double()
            for r in self.r_list:#不同r的列表
                start_time = time.time()
                print('aggregation degree: {}'.format(r))
                LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
                Z_layers = {}
                for im in range(len(patch_tokens_list)):
                    patch_tokens = [p.to(self.device) for p in patch_tokens_list[im]]
                    with torch.no_grad(), torch.cuda.amp.autocast():
                        features = LNAMD_r._embed(patch_tokens)#[B,1369,2,1024]
                        features /= features.norm(dim=-1, keepdim=True)
                        for l in range(len(self.features_list)):
                            # save the aggregated features
                            if str(l) not in Z_layers.keys():
                                Z_layers[str(l)] = []
                            Z_layers[str(l)].append(features[:, :, l, :])
                end_time = time.time()
                print('LNAMD-{}: {}ms per image'.format(r, (end_time-start_time)*1000/subset_num))

                # MSM
                anomaly_maps_l = torch.tensor([]).double()
                start_time = time.time()
                for l in Z_layers.keys():#Z_layers为字典，层数为键值数，值为20的列表
                    # different layers
                    Z = torch.cat(Z_layers[l], dim=0).to(self.device) # (83, 1369, 1024)
                    print('layer-{} mutual scoring...'.format(l))
                    anomaly_maps_msm = MSM(Z=Z, device=self.device, topmin_min=0, topmin_max=0.3)#(83,1369)
                    anomaly_maps_l = torch.cat((anomaly_maps_l, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)
                    torch.cuda.empty_cache()#(2,83,1369)
                anomaly_maps_l = torch.mean(anomaly_maps_l, 0)#(83,1369)
                anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_l.unsqueeze(0)), dim=0)#(r,83,1369)
                end_time = time.time()
                print('MSM: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            anomaly_maps_iter = torch.mean(anomaly_maps_r, 0).to(self.device)#取不同r的均值(83,1369)
            del anomaly_maps_r
            torch.cuda.empty_cache()

            # interpolate
            B, L = anomaly_maps_iter.shape
            H = int(np.sqrt(L))
            anomaly_maps_iter = F.interpolate(anomaly_maps_iter.view(B, 1, H, H),
                                        size=self.image_size, mode='bilinear', align_corners=True)
            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_iter.cpu()), dim=0)

        # save image features for optimizing classification
        # cls_save_path = os.path.join('./image_features/{}_{}.dat'.format(dataset, category))
        # with open(cls_save_path, 'wb') as f:
        #     pickle.dump([np.array(class_tokens)], f)
        end_time_all = time.time()
        print('MuSc: {}ms per image'.format((end_time_all-start_time_all)*1000/dataset_num))

        anomaly_maps = anomaly_maps.cpu().numpy()
        torch.cuda.empty_cache()

        B = anomaly_maps.shape[0]   # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)#(83,)
        # RsCIN
        if self.dataset == 'visa':
            k_score = [1, 8, 9]
        elif self.dataset == 'mvtec_ad':
            k_score = [1, 2, 3]
        else:
            k_score = [1, 2, 3]
        scores_cls = RsCIN(ac_score, class_tokens, k_list=k_score)#返回的仍然是83维数组，但经过了改进

        print('computing metrics...')
        pr_sp = np.array(scores_cls)#图片级别的异常分数
        gt_sp = np.array(gt_list)#真实标签0，1
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)#每张图片的每个像素点的真实标签
        pr_px = np.array(anomaly_maps)#每张图片的每个像素点的异常分数
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)#计算各项指标
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))

        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)
    
        return image_metric, pixel_metric


    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category,)
        #     auroc_sp, f1_sp, ap_sp = image_metric
        #     auroc_px, f1_px, ap_px, aupro = pixel_metric
        #     auroc_sp_ls.append(auroc_sp)
        #     f1_sp_ls.append(f1_sp)
        #     ap_sp_ls.append(ap_sp)
        #     auroc_px_ls.append(auroc_px)
        #     f1_px_ls.append(f1_px)
        #     ap_px_ls.append(ap_px)
        #     aupro_ls.append(aupro)
        # # mean
        # auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        # f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        # ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        # auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        # f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        # ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        # aupro_mean = sum(aupro_ls) / len(aupro_ls)

        # for i, category in enumerate(self.categories):
        #     print(category)
        #     print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
        #     print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        # print('mean')
        # print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        # print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        
        # # save in excel
        # if self.save_excel:
        #     workbook = Workbook()
        #     sheet = workbook.active
        #     sheet.title = "MuSc_results"
        #     sheet.cell(row=1,column=2,value='auroc_px')
        #     sheet.cell(row=1,column=3,value='f1_px')
        #     sheet.cell(row=1,column=4,value='ap_px')
        #     sheet.cell(row=1,column=5,value='aupro')
        #     sheet.cell(row=1,column=6,value='auroc_sp')
        #     sheet.cell(row=1,column=7,value='f1_sp')
        #     sheet.cell(row=1,column=8,value='ap_sp')
        #     for col_index in range(2):
        #         for row_index in range(len(self.categories)):
        #             if col_index == 0:
        #                 sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
        #             else:
        #                 sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
        #                 sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
        #                 sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
        #                 sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
        #                 sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
        #                 sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
        #                 sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
        #             if row_index == len(self.categories)-1:
        #                 if col_index == 0:
        #                     sheet.cell(row=row_index+3,column=col_index+1,value='mean')
        #                 else:
        #                     sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
        #                     sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
        #                     sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
        #                     sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
        #                     sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
        #                     sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
        #                     sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
        #     workbook.save(os.path.join(self.output_dir, 'results.xlsx'))


